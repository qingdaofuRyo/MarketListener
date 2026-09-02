"""Build the bundled TongdaXin/index-administrator Chinese-name catalogue."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


_DS_RECORD_SIZE = 106
_DS_RECORDS_OFFSET = 36
_DS_PREFIX_EXCHANGE = {
    10: "BASIC_FX",
    12: "GLOBAL_INDEX",
    16: "COMEX",
    17: "NYMEX",
    18: "CBOT",
    27: "HKEX",
    31: "HKEX",
    38: "TDX_MACRO",
    48: "HKEX",
    62: "CSI",
    69: "HUAZHENG",
    102: "CNI",
}


def _text_field(raw: bytes) -> str:
    return raw.split(b"\0", 1)[0].decode("gb18030", errors="replace").strip()


def load_tdx_ds_names(path: Path) -> dict[str, str]:
    """Read prefix-aware names from TongdaXin ``ds_stk.dat`` records."""
    raw = path.read_bytes()
    output: dict[str, str] = {}
    for offset in range(_DS_RECORDS_OFFSET, len(raw) - _DS_RECORD_SIZE + 1, _DS_RECORD_SIZE):
        prefix = raw[offset - 4]
        exchange = _DS_PREFIX_EXCHANGE.get(prefix)
        if not exchange:
            continue
        symbol = _text_field(raw[offset : offset + 24])
        name = _text_field(raw[offset + 23 : offset + 65])
        if symbol and name and symbol.casefold() != name.casefold():
            output[f"{exchange}.{symbol.upper()}"] = name
    return output


def load_index_workbook(path: Path, exchange: str) -> dict[str, str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    code_index = headers.index("指数代码")
    name_index = headers.index("指数简称")
    output: dict[str, str] = {}
    for row in rows:
        symbol = str(row[code_index] or "").strip().upper()
        name = str(row[name_index] or "").strip()
        if symbol and name and symbol.casefold() != name.casefold():
            output[f"{exchange}.{symbol}"] = name
    return output


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--existing", type=Path, required=True)
    parser.add_argument("--tdx-ds-stock", type=Path, required=True)
    parser.add_argument("--csi", type=Path, required=True)
    parser.add_argument("--cni", type=Path, required=True)
    parser.add_argument("--huazheng", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    payload: dict[str, Any] = json.loads(args.existing.read_text(encoding="utf-8"))
    names = {str(key): str(value) for key, value in payload.items()}
    source_counts: dict[str, int] = {}
    sources = (
        ("tdx_ds", load_tdx_ds_names(args.tdx_ds_stock)),
        ("csi", load_index_workbook(args.csi, "CSI")),
        ("cni", load_index_workbook(args.cni, "CNI")),
        ("huazheng", load_index_workbook(args.huazheng, "HUAZHENG")),
    )
    for source, values in sources:
        names.update(values)
        source_counts[source] = len(values)
    args.output.write_text(
        json.dumps(dict(sorted(names.items())), ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(json.dumps({"total": len(names), "sources": source_counts}, ensure_ascii=False))


if __name__ == "__main__":
    main()
