"""真实数据采集会话：抓取用户清单中的行情/宏观/特色数据并落库。

每次 ``run_fetch_session`` 会把每个数据集任务的结果写入 Silver 分区或
Gold 指标表，并把任务级进展写入 ``data_root/control_summary.json``，
供 HTML 控制中心展示“抓取了哪些类型、进展如何”。
"""

from __future__ import annotations

import concurrent.futures
import csv
from io import BytesIO, StringIO
import json
import os
import re
import time
import urllib.error
import urllib.parse
import urllib.request
from contextlib import contextmanager
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Callable, Iterator, Mapping, Sequence
from uuid import uuid4

from openpyxl import load_workbook

from .futures_dashboard import (
    MemberPositionRow,
    build_open_interest_leaderboard,
    compute_futures_breadth,
)
from .futures_member_positions import (
    MemberPositionRank,
    collect_exchange_member_position_ranks,
)
from .macro_series import MacroPoint, derive_series, macro_series_index, normalise_macro_point
from .storage import MarketStore, PartitionKey

_TDX_SERVERS: tuple[tuple[str, int], ...] = (
    ("60.191.117.167", 7709),
    ("218.75.126.9", 7709),
    ("115.238.56.198", 7709),
)
_BEA_TRADE_PAGE_URL = "https://www.bea.gov/data/intl-trade-investment/international-trade-goods-and-services"
_BEA_TRADE_XLSX_RE = re.compile(r'href=["\'](?P<url>[^"\']*trad\d{4}-time-series\.xlsx)["\']', re.IGNORECASE)
_BEA_MONTH_RE = re.compile(r"^(?P<year>\d{4})\s+(?P<month>[A-Za-z]{3})")
_BEA_GDP_PAGE_URL = "https://www.bea.gov/data/gdp/gross-domestic-product"
_BEA_NIPA_QUARTERLY_URL = "https://apps.bea.gov/national/Release/TXT/NipaDataQ.txt"
_BEA_GDP_RELEASE_RE = re.compile(
    r"(?P<stage>Advance|Second Estimate|Third Estimate).*?(?P<quarter>[1-4])(?:st|nd|rd|th) Quarter (?P<year>\d{4})",
    re.IGNORECASE,
)
_BEA_NIPA_CORE_PCE_CODE = "DPCCRG"
_MONTHS_BY_ABBREVIATION = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


@dataclass
class BarWrite:
    """一份待写 Silver 分区的 K 线数据。"""

    key: PartitionKey
    bars: list[dict[str, Any]]
    data_cutoff: str


@dataclass
class PersistPlan:
    """一个采集任务的落库内容。"""

    bar_writes: list[BarWrite] = field(default_factory=list)
    gold_metrics: list[dict[str, Any]] = field(default_factory=list)
    member_position_ranks: list[dict[str, Any]] = field(default_factory=list)
    member_position_coverage: list[dict[str, Any]] = field(default_factory=list)


@dataclass(frozen=True)
class CollectionTask:
    """一个可独立抓取的数据集任务。"""

    dataset_id: str
    dataset_name: str
    source: str
    fn: Callable[[], "CollectionTaskResult | list[CollectionTaskResult]"]


@dataclass
class CollectionTaskResult:
    """一个数据集任务的抓取结果与落库计划。"""

    dataset_id: str
    dataset_name: str
    source: str
    status: str = "PASS"
    rows: int = 0
    detail: str = ""
    error: str | None = None
    started_at: str = ""
    completed_at: str = ""
    persist: PersistPlan = field(default_factory=PersistPlan)

    def to_dict(self) -> dict[str, Any]:
        return {
            "dataset_id": self.dataset_id,
            "dataset_name": self.dataset_name,
            "source": self.source,
            "status": self.status,
            "rows": self.rows,
            "detail": self.detail,
            "error": self.error,
            "started_at": self.started_at,
            "completed_at": self.completed_at,
        }


def run_fetch_session(
    data_root: Path,
    *,
    tasks: Sequence[CollectionTask] | None = None,
    max_workers: int = 4,
    task_timeout_seconds: float = 90.0,
    limit_futures: int = 15,
    limit_cn_stocks: int = 5,
    now: datetime | None = None,
) -> dict[str, Any]:
    """运行一次真实数据采集会话，落库并生成 control_summary.json。"""

    data_root = Path(data_root)
    if max_workers <= 0:
        raise ValueError("max_workers must be positive")
    if task_timeout_seconds <= 0:
        raise ValueError("task_timeout_seconds must be positive")
    started_at = (now or datetime.now(timezone.utc)).isoformat(timespec="seconds")
    if tasks is None:
        tasks = _build_tasks(limit_futures=limit_futures, limit_cn_stocks=limit_cn_stocks)
    results = _run_tasks(tasks, max_workers=max_workers, task_timeout_seconds=task_timeout_seconds)
    results.sort(key=lambda result: (result.dataset_id, result.started_at))
    _persist_results(data_root, results)

    statuses = [result.status for result in results]
    if not statuses:
        session_status = "PASS"
    elif all(status == "PASS" for status in statuses):
        session_status = "PASS"
    elif all(status == "BLOCKED" for status in statuses):
        session_status = "BLOCKED"
    else:
        session_status = "PARTIAL_FAILURE"
    summary = {
        "session_id": f"session-{uuid4().hex[:12]}",
        "started_at": started_at,
        "completed_at": _now(),
        "status": session_status,
        "task_count": len(results),
        "passed": statuses.count("PASS"),
        "partial_failure": statuses.count("PARTIAL_FAILURE"),
        "failed": statuses.count("FAILED"),
        "blocked": statuses.count("BLOCKED"),
        "total_rows": sum(int(result.rows or 0) for result in results),
        "tasks": [result.to_dict() for result in results],
    }
    target = data_root / "control_summary.json"
    target.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return summary


def _build_tasks(*, limit_futures: int, limit_cn_stocks: int) -> list[CollectionTask]:
    return [
        CollectionTask(
            "CN_STOCK_BAR",
            "A股个股K线（样本）",
            "pytdx",
            lambda: _collect_cn_stock_bars(limit_cn_stocks),
        ),
        CollectionTask("CN_INDEX_BAR", "A股指数K线", "pytdx", _collect_cn_index_bars),
        CollectionTask("CN_ETF_BAR", "A股ETF K线", "pytdx", _collect_cn_etf_bars),
        CollectionTask("HK_STOCK_BAR", "港股个股K线", "akshare", _collect_hk_stock_bars),
        CollectionTask("HK_INDEX_BAR", "港股指数K线", "akshare", _collect_hk_index_bars),
        CollectionTask("GLOBAL_INDEX_BAR", "全球指数K线", "akshare", _collect_global_indices),
        CollectionTask(
            "FUTURE_MAIN_BAR",
            "国内期货主力K线",
            "akshare",
            lambda: _collect_futures_main(limit_futures),
        ),
        CollectionTask("FUTURE_INDEX_BAR", "国内商品指数K线", "akshare", _collect_futures_index),
        CollectionTask("FUTURES_OI_LEADERBOARD", "期货持仓龙虎榜", "akshare", _collect_oi_leaderboard),
        CollectionTask("CN_MARGIN", "沪深京融资融券", "akshare", _collect_margin),
        CollectionTask("MACRO_SERIES", "宏观数据序列", "akshare", _collect_macro),
        CollectionTask("A_SHARE_BREADTH", "A股涨跌/市值快照", "akshare", _collect_a_share_breadth),
        CollectionTask("HSGT_FLOW", "北向南向资金", "akshare", _collect_hsgt),
        CollectionTask("CN_ZT_POOL", "涨停/跌停池与连板高度", "akshare", _collect_zt_pool),
        CollectionTask("FUTURE_GLOBAL_BAR", "外盘期货/金银比/金油比", "akshare", _collect_foreign_futures),
        CollectionTask("CRYPTO_BAR", "加密货币K线", "binance", _collect_crypto),
        CollectionTask("USD_INDEX_VIX", "美元指数/VIX", "eastmoney+cboe", _collect_usd_vix),
    ]


def _run_tasks(
    tasks: Sequence[CollectionTask],
    *,
    max_workers: int,
    task_timeout_seconds: float,
) -> list[CollectionTaskResult]:
    results: list[CollectionTaskResult] = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers, thread_name_prefix="collector") as pool:
        futures = {pool.submit(_execute_task, task): task for task in tasks}
        for future in concurrent.futures.as_completed(futures):
            task = futures[future]
            try:
                outcome = future.result(timeout=task_timeout_seconds)
            except Exception as error:
                outcome = _failed_result(task, error)
            if isinstance(outcome, list):
                results.extend(outcome)
            else:
                results.append(outcome)
    return results


def _execute_task(task: CollectionTask) -> CollectionTaskResult | list[CollectionTaskResult]:
    started_at = _now()
    try:
        outcome = task.fn()
    except Exception as error:
        return _failed_result(task, error, started_at=started_at)
    items = outcome if isinstance(outcome, list) else [outcome]
    for item in items:
        if not item.started_at:
            item.started_at = started_at
        if not item.completed_at:
            item.completed_at = _now()
    return outcome


def _failed_result(
    task: CollectionTask,
    error: Exception,
    *,
    started_at: str | None = None,
) -> CollectionTaskResult:
    detail = f"{type(error).__name__}: {error}"
    return CollectionTaskResult(
        dataset_id=task.dataset_id,
        dataset_name=task.dataset_name,
        source=task.source,
        status="FAILED",
        rows=0,
        detail=detail[:500],
        error=str(error)[:500],
        started_at=started_at or _now(),
        completed_at=_now(),
    )


def _persist_results(data_root: Path, results: Sequence[CollectionTaskResult]) -> None:
    store = MarketStore(data_root)
    try:
        store.register_default_datasets()
        for result in results:
            plan = result.persist
            if (
                not plan.bar_writes
                and not plan.gold_metrics
                and not plan.member_position_ranks
                and not plan.member_position_coverage
            ):
                continue
            run_id = store.begin_run(f"collector:{result.source}")
            try:
                for write in plan.bar_writes:
                    store.write_silver_bars(write.key, write.bars, write.data_cutoff, run_id)
                if plan.gold_metrics:
                    store.upsert_gold_metrics(plan.gold_metrics)
                if plan.member_position_ranks:
                    store.upsert_futures_member_position_ranks(plan.member_position_ranks)
                if plan.member_position_coverage:
                    store.upsert_futures_member_position_coverage(plan.member_position_coverage)
                if result.status == "PASS":
                    store.finish_run(run_id, "COMPLETE", result.detail)
                else:
                    store.finish_run(run_id, "PARTIAL_FAILURE", result.detail)
            except Exception as error:
                message = f"persist failed: {type(error).__name__}: {error}"
                store.finish_run(run_id, "FAILED", message[:500])
                result.status = "FAILED"
                result.error = message[:500]
                result.detail = message[:500]
    finally:
        store.close()


def _result(
    dataset_id: str,
    dataset_name: str,
    source: str,
    status: str,
    rows: int,
    detail: str,
    error: str | None,
    plan: PersistPlan,
) -> CollectionTaskResult:
    return CollectionTaskResult(
        dataset_id=dataset_id,
        dataset_name=dataset_name,
        source=source,
        status=status,
        rows=rows,
        detail=detail,
        error=error,
        persist=plan,
    )


def _now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _today_compact() -> str:
    return date.today().isoformat().replace("-", "")


def _last_trading_day_compact() -> str:
    day = date.today()
    while day.weekday() >= 5:
        day -= timedelta(days=1)
    return day.isoformat().replace("-", "")


def _num(value: Any, default: float | None = None) -> float | None:
    if value is None:
        return default
    try:
        number = float(value)
    except (TypeError, ValueError):
        return default
    return number if number == number else default


def _ak() -> Any:
    import akshare

    return _DirectApi(akshare)


@contextmanager
def _direct_network() -> Iterator[None]:
    """临时绕过系统代理，解决本机 MITM 代理导致 akshare HTTPS 证书校验失败的问题。"""

    previous = os.environ.get("NO_PROXY")
    os.environ["NO_PROXY"] = "*"
    try:
        yield
    finally:
        if previous is None:
            os.environ.pop("NO_PROXY", None)
        else:
            os.environ["NO_PROXY"] = previous


class _DirectApi:
    """akshare 委托包装：每次调用都直连网络，避免系统代理拦截 HTTPS。"""

    def __init__(self, target: Any) -> None:
        self._target = target

    def __getattr__(self, name: str) -> Any:
        attr = getattr(self._target, name)
        if not callable(attr):
            return attr

        def wrapper(*args: Any, **kwargs: Any) -> Any:
            with _direct_network():
                return attr(*args, **kwargs)

        return wrapper


def _latest_by_date(
    records: Sequence[Mapping[str, Any]], date_keys: Sequence[str]
) -> tuple[dict[str, Any] | None, str]:
    """按日期字段取最新一条记录（兼容升序/降序与 YYYYMMDD/YYYY-MM-DD）。"""

    latest: dict[str, Any] | None = None
    latest_day = ""
    for row in records:
        for key in date_keys:
            raw = row.get(key)
            if raw is None:
                continue
            day = str(raw).replace("-", "").replace("/", "")
            if day > latest_day:
                latest_day = day
                latest = dict(row)
            break
    return latest, latest_day


def _daily_bar(
    instrument_id: str,
    name: str,
    trading_date: Any,
    *,
    open_: Any,
    high: Any,
    low: Any,
    close: Any,
    volume: Any = None,
    amount: Any = None,
    open_interest: Any = None,
    pct_change: Any = None,
    period: str = "1d",
    source: str = "akshare",
    currency: str | None = None,
    adjustment: str = "",
) -> dict[str, Any]:
    trading_date = str(trading_date)[:10]
    close = _num(close, 0.0) or 0.0
    open_price = _num(open_, close) or close
    high_price = _num(high, max(open_price, close))
    low_price = _num(low, min(open_price, close))
    amplitude = (high_price - low_price) / low_price * 100 if low_price else None
    bar_open_time = f"{trading_date}T00:00:00+08:00"
    return {
        "instrument_id": instrument_id,
        "symbol": instrument_id.rsplit(".", 1)[-1],
        "name": name,
        "trading_date": trading_date,
        "bar_start": bar_open_time,
        "bar_end": f"{trading_date}T23:59:59+08:00",
        "bar_open_time": bar_open_time,
        "period": period,
        "open": open_price,
        "high": high_price,
        "low": low_price,
        "close": close,
        "pct_change": _num(pct_change),
        "amplitude": amplitude,
        "volume": _num(volume),
        "amount": _num(amount),
        "open_interest": _num(open_interest),
        "currency": currency,
        "adjustment": adjustment,
        "source": source,
        "source_period": period,
        "fetched_at": _now(),
        "data_version": "1",
        "quality_status": "PASS",
    }


def _bar_partitions(
    market: str,
    asset_type: str,
    period: str,
    bars: Sequence[Mapping[str, Any]],
    prefix: str,
) -> list[BarWrite]:
    by_year: dict[int, list[dict[str, Any]]] = {}
    for bar in bars:
        raw = str(bar.get("bar_open_time") or bar.get("trading_date") or "")[:10]
        try:
            year = int(raw[:4])
        except ValueError:
            continue
        by_year.setdefault(year, []).append(dict(bar))
    writes: list[BarWrite] = []
    current_year = date.today().year
    for year in sorted(by_year):
        if year < 1990 or year > current_year + 1:
            continue
        grouped = by_year[year]
        grouped.sort(key=lambda bar: str(bar.get("bar_open_time") or ""))
        cutoff = max(str(bar.get("bar_open_time") or "") for bar in grouped) or f"{year}-01-01"
        writes.append(
            BarWrite(
                PartitionKey(market, asset_type, period, year, f"{prefix}-{year}"),
                grouped,
                cutoff,
            )
        )
    return writes


def _metric(
    series_id: str,
    instrument_id: str,
    trading_date: str,
    period: str,
    metric_name: str,
    value: Any,
    definition: str,
    calculation_method: str,
    *,
    metric_key: str = "value",
    timestamp: str | None = None,
) -> dict[str, Any]:
    number = _num(value)
    if number is None:
        raise ValueError(f"metric value must be numeric: {metric_name}={value!r}")
    return {
        "metric_id": f"{series_id}:{instrument_id}:{trading_date}:{period}:{metric_key}",
        "instrument_id": instrument_id,
        "trading_date": str(trading_date),
        "period": period,
        "metric_name": metric_name,
        "value": number,
        "definition": definition,
        "calculation_method": calculation_method,
        "timestamp": timestamp or _now(),
    }


def _macro_metrics(points: Sequence[MacroPoint]) -> list[dict[str, Any]]:
    return [
        _metric(
            point.series_id,
            point.series_id,
            point.available_time,
            point.frequency,
            point.name,
            point.value,
            point.definition,
            point.calculation_method,
            metric_key="value",
            timestamp=point.fetched_at,
        )
        for point in points
    ]


def _parse_cn_month(value: Any) -> str:
    text = (
        str(value)
        .strip()
        .replace("年", "-")
        .replace("月份", "")
        .replace("月", "")
        .replace("/", "-")
        .replace(".", "-")
    )
    parts = text.split("-")
    if len(parts) >= 2:
        try:
            return f"{int(parts[0]):04d}-{int(parts[1]):02d}"
        except ValueError:
            pass
    return str(value)[:7]


def _discover_bea_trade_xlsx_url(page_html: str) -> str:
    """Find the current BEA time-series workbook without hard-coding a release month."""

    match = _BEA_TRADE_XLSX_RE.search(page_html)
    if match is None:
        raise ValueError("BEA trade time-series workbook link was not found")
    return urllib.parse.urljoin(_BEA_TRADE_PAGE_URL, match.group("url"))


def _parse_bea_trade_imports_xlsx(workbook_bytes: bytes) -> list[tuple[str, float]]:
    """Read BEA Table 1's seasonally adjusted monthly total imports.

    The official workbook distinguishes the total goods-and-services import
    column from goods-only and services-only columns.  Validate both header
    rows before consuming numeric values so a spreadsheet layout change cannot
    silently substitute an adjacent measure.
    """

    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    try:
        if "Table 1" not in workbook.sheetnames:
            raise ValueError("BEA trade workbook is missing Table 1")
        sheet = workbook["Table 1"]
        rows = sheet.iter_rows(values_only=True)
        header: tuple[Any, ...] | None = None
        subheader: tuple[Any, ...] | None = None
        import_total_index: int | None = None
        result: list[tuple[str, float]] = []
        for row in rows:
            cells = tuple(row)
            if header is None:
                imports_heading_index = next(
                    (
                        index
                        for index, cell in enumerate(cells)
                        if isinstance(cell, str) and cell.strip() == "Imports"
                    ),
                    None,
                )
                if imports_heading_index is None or imports_heading_index == 0:
                    continue
                # In the official sheet the merged Imports heading is stored
                # one cell to the right of its three-column group.  Its left
                # edge is the Total column used for the requested series.
                import_total_index = imports_heading_index - 1
                header = cells
                continue
            if header is not None and subheader is None:
                if import_total_index is not None and len(cells) > import_total_index and cells[import_total_index] == "Total":
                    subheader = cells
                continue
            if subheader is None or not cells or not isinstance(cells[0], str):
                continue
            match = _BEA_MONTH_RE.match(cells[0].strip())
            if match is None:
                continue
            month_number = _MONTHS_BY_ABBREVIATION.get(match.group("month").lower())
            total_imports = _num(cells[import_total_index] if import_total_index is not None and len(cells) > import_total_index else None)
            if month_number is not None and total_imports is not None:
                result.append((f"{match.group('year')}-{month_number:02d}", total_imports))
        if header is None or subheader is None:
            raise ValueError("BEA trade workbook headers do not identify total imports")
        if not result:
            raise ValueError("BEA trade workbook contains no monthly total-import observations")
        return result
    finally:
        workbook.close()


def _fetch_bea_us_imports_rows() -> list[tuple[str, float]]:
    """Fetch official BEA seasonally adjusted goods-and-services imports."""

    request_headers = {"User-Agent": "MarketListener/0.1 (+local research terminal)"}
    with _direct_network():
        page_request = urllib.request.Request(_BEA_TRADE_PAGE_URL, headers=request_headers)
        with urllib.request.urlopen(page_request, timeout=30) as response:  # noqa: S310 - fixed official BEA URL
            workbook_url = _discover_bea_trade_xlsx_url(response.read().decode("utf-8", errors="replace"))
        workbook_request = urllib.request.Request(workbook_url, headers=request_headers)
        with urllib.request.urlopen(workbook_request, timeout=30) as response:  # noqa: S310 - parsed from official BEA page
            return _parse_bea_trade_imports_xlsx(response.read())


def _parse_bea_gdp_final_cutoff(page_html: str) -> tuple[int, int]:
    """Return the latest quarter with a BEA GDP third (final) estimate."""

    match = _BEA_GDP_RELEASE_RE.search(page_html)
    if match is None:
        raise ValueError("BEA GDP release stage and quarter were not found")
    year = int(match.group("year"))
    quarter = int(match.group("quarter"))
    if match.group("stage").lower() == "third estimate":
        return year, quarter
    if quarter == 1:
        return year - 1, 4
    return year, quarter - 1


def _parse_bea_core_pce_final_rows(
    nipa_text: str,
    *,
    final_through: tuple[int, int],
) -> list[tuple[str, float]]:
    """Calculate core-PCE annualized QoQ changes through BEA's final cutoff.

    ``DPCCRG`` is the quarterly NIPA 2.3.4 index for PCE excluding food and
    energy.  The official GDP release explains that quarterly percentage
    changes are annual rates; derive that rate from the consecutive published
    seasonal index levels and discard the in-progress quarter until a third
    estimate exists.
    """

    levels: list[tuple[int, int, float]] = []
    for row in csv.reader(StringIO(nipa_text)):
        if len(row) != 3 or row[0] != _BEA_NIPA_CORE_PCE_CODE:
            continue
        period_match = re.fullmatch(r"(?P<year>\d{4})Q(?P<quarter>[1-4])", row[1])
        if period_match is None:
            continue
        try:
            level = float(row[2].replace(",", ""))
        except ValueError:
            continue
        if level > 0:
            levels.append((int(period_match.group("year")), int(period_match.group("quarter")), level))
    levels.sort()
    if len(levels) < 2:
        raise ValueError("BEA NIPA data contains too few core-PCE index observations")

    result: list[tuple[str, float]] = []
    previous_year, previous_quarter, previous_level = levels[0]
    for year, quarter, level in levels[1:]:
        expected_previous = (year, quarter - 1) if quarter > 1 else (year - 1, 4)
        if (previous_year, previous_quarter) == expected_previous and (year, quarter) <= final_through:
            annualized_change = ((level / previous_level) ** 4 - 1.0) * 100.0
            result.append((f"{year}-Q{quarter}", annualized_change))
        previous_year, previous_quarter, previous_level = year, quarter, level
    if not result:
        raise ValueError("BEA NIPA data contains no core-PCE observations through the final cutoff")
    return result


def _fetch_bea_core_pce_final_rows() -> list[tuple[str, float]]:
    """Fetch the BEA core-PCE SAAR series only through the latest final GDP quarter."""

    request_headers = {"User-Agent": "MarketListener/0.1 (+local research terminal)"}
    with _direct_network():
        gdp_request = urllib.request.Request(_BEA_GDP_PAGE_URL, headers=request_headers)
        with urllib.request.urlopen(gdp_request, timeout=30) as response:  # noqa: S310 - fixed official BEA URL
            final_through = _parse_bea_gdp_final_cutoff(response.read().decode("utf-8", errors="replace"))
        nipa_request = urllib.request.Request(_BEA_NIPA_QUARTERLY_URL, headers=request_headers)
        with urllib.request.urlopen(nipa_request, timeout=60) as response:  # noqa: S310 - fixed official BEA URL
            return _parse_bea_core_pce_final_rows(
                response.read().decode("utf-8", errors="replace"),
                final_through=final_through,
            )


def _tdx_bars(
    period: str,
    specs: Sequence[tuple[str, int, str]],
    count: int,
    *,
    asset_type: str = "STOCK",
) -> tuple[list[dict[str, Any]], list[str]]:
    from pytdx.hq import TdxHq_API

    category = {"1d": 4, "30m": 2}[period]
    all_errors: list[str] = []
    for host, port in _TDX_SERVERS:
        api = TdxHq_API()
        server_errors: list[str] = []
        try:
            if not api.connect(host, port, time_out=8):
                all_errors.append(f"{host}:{port} refused")
                continue
            bars: list[dict[str, Any]] = []
            for code, market, name in specs:
                try:
                    fetch = api.get_index_bars if asset_type == "INDEX" else api.get_security_bars
                    rows = fetch(category, market, code, 0, count) or []
                except Exception as error:
                    server_errors.append(f"{code}:{type(error).__name__}")
                    continue
                if not rows:
                    server_errors.append(f"{code}:empty")
                    continue
                exchange = "SSE" if market == 1 else "SZSE"
                for row in rows:
                    bars.append(
                        _tdx_bar(
                            row,
                            f"CN.{exchange}.{asset_type}.{code}",
                            name,
                            period,
                        )
                    )
            if bars:
                return bars, server_errors
            all_errors.append(f"{host}:{port}:all-empty")
        except Exception as error:
            all_errors.append(f"{host}:{port}:{type(error).__name__}")
        finally:
            try:
                api.disconnect()
            except Exception:
                pass
    return [], all_errors


def _tdx_bar(row: Mapping[str, Any], instrument_id: str, name: str, period: str) -> dict[str, Any]:
    year, month, day = int(row["year"]), int(row["month"]), int(row["day"])
    if period == "1d":
        bar_open_time = f"{year:04d}-{month:02d}-{day:02d}T00:00:00+08:00"
    else:
        hour = int(row.get("hour") or 0)
        minute = int(row.get("minute") or 0)
        bar_open_time = f"{year:04d}-{month:02d}-{day:02d}T{hour:02d}:{minute:02d}:00+08:00"
    return _daily_bar(
        instrument_id,
        name,
        bar_open_time[:10],
        open_=row.get("open"),
        high=row.get("high"),
        low=row.get("low"),
        close=row.get("close"),
        volume=row.get("vol"),
        amount=row.get("amount"),
        open_interest=row.get("position"),
        period=period,
        source="pytdx",
    )


def _collect_cn_stock_bars(max_stocks: int) -> CollectionTaskResult:
    specs = [
        ("600519", 1, "贵州茅台"),
        ("600036", 1, "招商银行"),
        ("000001", 0, "平安银行"),
        ("300750", 0, "宁德时代"),
        ("000858", 0, "五粮液"),
    ][:max_stocks]
    if not specs:
        raise ValueError("at least one CN stock is required")
    bars_1d, errors_1d = _tdx_bars("1d", specs, 300, asset_type="STOCK")
    bars_30m, errors_30m = _tdx_bars("30m", specs[:1], 120, asset_type="STOCK")
    plan = PersistPlan()
    if bars_1d:
        plan.bar_writes.extend(_bar_partitions("CN", "STOCK", "1d", bars_1d, "CN-STOCK-1d"))
    if bars_30m:
        plan.bar_writes.extend(_bar_partitions("CN", "STOCK", "30m", bars_30m, "CN-STOCK-30m"))
    errors = list(dict.fromkeys(errors_1d + errors_30m))
    rows = len(bars_1d) + len(bars_30m)
    return _result(
        "CN_STOCK_BAR",
        "A股个股K线（样本）",
        "pytdx",
        "PASS" if not errors and rows else "PARTIAL_FAILURE",
        rows,
        f"1d={len(bars_1d)} 30m={len(bars_30m)}；错误={'; '.join(errors) or '无'}",
        "; ".join(errors) or None,
        plan,
    )


def _collect_cn_index_bars() -> CollectionTaskResult:
    specs = [
        ("000001", 1, "上证指数"),
        ("000300", 1, "沪深300"),
        ("399001", 0, "深证成指"),
        ("399006", 0, "创业板指"),
    ]
    bars, errors = _tdx_bars("1d", specs, 300, asset_type="INDEX")
    plan = PersistPlan()
    if bars:
        plan.bar_writes.extend(_bar_partitions("CN", "INDEX", "1d", bars, "CN-INDEX-1d"))
    return _result(
        "CN_INDEX_BAR",
        "A股指数K线",
        "pytdx",
        "PASS" if not errors and bars else "PARTIAL_FAILURE",
        len(bars),
        f"指数1d={len(bars)}；错误={'; '.join(errors) or '无'}",
        "; ".join(errors) or None,
        plan,
    )


def _collect_cn_etf_bars() -> CollectionTaskResult:
    specs = [
        ("510300", 1, "沪深300ETF"),
        ("510500", 1, "中证500ETF"),
        ("512100", 1, "中证1000ETF"),
        ("159915", 0, "创业板ETF"),
    ]
    bars, errors = _tdx_bars("1d", specs, 300, asset_type="ETF")
    plan = PersistPlan()
    if bars:
        plan.bar_writes.extend(_bar_partitions("CN", "ETF", "1d", bars, "CN-ETF-1d"))
    return _result(
        "CN_ETF_BAR",
        "A股ETF K线",
        "pytdx",
        "PASS" if not errors and bars else "PARTIAL_FAILURE",
        len(bars),
        f"ETF 1d={len(bars)}；错误={'; '.join(errors) or '无'}",
        "; ".join(errors) or None,
        plan,
    )


def _collect_hk_stock_bars() -> CollectionTaskResult:
    api = _ak()
    specs = [("00700", "腾讯控股"), ("09988", "阿里巴巴-W"), ("03690", "美团-W"), ("00981", "中芯国际")]
    bars: list[dict[str, Any]] = []
    errors: list[str] = []
    for code, name in specs:
        try:
            frame = api.stock_hk_daily(symbol=code, adjust="qfq")
        except Exception as error:
            errors.append(f"{code}:{type(error).__name__}")
            continue
        for row in frame.to_dict(orient="records"):
            bars.append(
                _daily_bar(
                    f"HK.HKEX.STOCK.{code}",
                    name,
                    row["date"],
                    open_=row.get("open"),
                    high=row.get("high"),
                    low=row.get("low"),
                    close=row.get("close"),
                    volume=row.get("volume"),
                    amount=row.get("amount"),
                    currency="HKD",
                    source="akshare-stock_hk_daily",
                )
            )
    plan = PersistPlan()
    if bars:
        plan.bar_writes.extend(_bar_partitions("HK", "STOCK", "1d", bars, "HK-STOCK-1d"))
    return _result(
        "HK_STOCK_BAR",
        "港股个股K线",
        "akshare",
        "PASS" if not errors and bars else "PARTIAL_FAILURE",
        len(bars),
        f"4只港股日报={len(bars)}；错误={'; '.join(errors) or '无'}",
        "; ".join(errors) or None,
        plan,
    )


def _collect_hk_index_bars() -> CollectionTaskResult:
    api = _ak()
    specs = [("HSI", "恒生指数"), ("HSCEI", "恒生国企指数"), ("HSTECH", "恒生科技指数")]
    bars: list[dict[str, Any]] = []
    errors: list[str] = []
    for symbol, name in specs:
        try:
            frame = api.stock_hk_index_daily_sina(symbol=symbol)
        except Exception as error:
            errors.append(f"{symbol}:{type(error).__name__}")
            continue
        for row in frame.to_dict(orient="records"):
            bars.append(
                _daily_bar(
                    f"HK.HKEX.INDEX.{symbol}",
                    name,
                    row["date"],
                    open_=row.get("open"),
                    high=row.get("high"),
                    low=row.get("low"),
                    close=row.get("close"),
                    volume=row.get("volume"),
                    amount=row.get("amount"),
                    currency="HKD",
                    source="akshare-stock_hk_index_daily_sina",
                )
            )
    plan = PersistPlan()
    if bars:
        plan.bar_writes.extend(_bar_partitions("HK", "INDEX", "1d", bars, "HK-INDEX-1d"))
    return _result(
        "HK_INDEX_BAR",
        "港股指数K线",
        "akshare",
        "PASS" if not errors and bars else "PARTIAL_FAILURE",
        len(bars),
        f"港股指数日线={len(bars)}；错误={'; '.join(errors) or '无'}",
        "; ".join(errors) or None,
        plan,
    )


def _collect_global_indices() -> CollectionTaskResult:
    api = _ak()
    attempts = [
        ("index_us_stock_sina", {"symbol": ".INX"}, "US.SPX.INDEX.S&P500", "标普500"),
        ("index_us_stock_sina", {"symbol": ".DJI"}, "US.DJIA.INDEX.DOW", "道琼斯工业"),
        ("index_us_stock_sina", {"symbol": ".IXIC"}, "US.NASDAQ.INDEX.COMP", "纳斯达克综合"),
        ("index_global_hist_sina", {"symbol": "\u65e5\u7ecf225\u6307\u6570"}, "JP.N225.INDEX.NIKKEI225", "日经225"),
        ("index_global_hist_sina", {"symbol": "\u82f1\u56fd\u5bcc\u65f6100\u6307\u6570"}, "UK.FTSE.INDEX.FTSE100", "英国富时100"),
        ("index_global_hist_sina", {"symbol": "\u5fb7\u56fdDAX 30\u79cd\u80a1\u4ef7\u6307\u6570"}, "DE.DAX.INDEX.DAX30", "德国DAX30"),
    ]
    bars: list[dict[str, Any]] = []
    errors: list[str] = []
    for function_name, kwargs, instrument_id, name in attempts:
        function = getattr(api, function_name, None)
        if function is None:
            errors.append(f"{function_name}:missing")
            continue
        try:
            frame = function(**kwargs)
        except Exception as error:
            errors.append(f"{name}:{type(error).__name__}")
            continue
        for row in frame.to_dict(orient="records"):
            bars.append(
                _daily_bar(
                    instrument_id,
                    name,
                    row["date"],
                    open_=row.get("open"),
                    high=row.get("high"),
                    low=row.get("low"),
                    close=row.get("close"),
                    volume=row.get("volume"),
                    amount=row.get("amount"),
                    currency="USD" if "US." in instrument_id else None,
                    source="akshare-global-index",
                )
            )
    plan = PersistPlan()
    if bars:
        plan.bar_writes.extend(_bar_partitions("GLOBAL", "INDEX", "1d", bars, "GLOBAL-INDEX-1d"))
    return _result(
        "GLOBAL_INDEX_BAR",
        "全球指数K线",
        "akshare",
        "PASS" if not errors and bars else "PARTIAL_FAILURE",
        len(bars),
        f"全球指数日线={len(bars)}；错误={'; '.join(errors) or '无'}",
        "; ".join(errors) or None,
        plan,
    )


def _collect_futures_main(limit: int) -> list[CollectionTaskResult]:
    api = _ak()
    frame = api.futures_display_main_sina()
    selected = frame.to_dict(orient="records")
    if limit and limit > 0:
        selected = selected[:limit]
    today = _today_compact()
    bars: list[dict[str, Any]] = []
    errors: list[str] = []
    for row in selected:
        symbol = str(row["symbol"])
        exchange = str(row.get("exchange") or "").upper()
        name = str(row.get("name") or symbol)
        try:
            daily = api.futures_main_sina(symbol=symbol, start_date="20240101", end_date=today)
        except Exception as error:
            errors.append(f"{symbol}:{type(error).__name__}")
            continue
        for bar in daily.to_dict(orient="records"):
            bars.append(
                _daily_bar(
                    f"CN.{exchange}.FUTURE.{symbol}",
                    name,
                    bar["\u65e5\u671f"],
                    open_=bar.get("\u5f00\u76d8\u4ef7"),
                    high=bar.get("\u6700\u9ad8\u4ef7"),
                    low=bar.get("\u6700\u4f4e\u4ef7"),
                    close=bar.get("\u6536\u76d8\u4ef7"),
                    volume=bar.get("\u6210\u4ea4\u91cf"),
                    amount=None,
                    open_interest=bar.get("\u6301\u4ed3\u91cf"),
                    source="akshare-futures_main_sina",
                )
            )
    plan = PersistPlan()
    if bars:
        plan.bar_writes.extend(_bar_partitions("CN", "FUTURE", "1d", bars, "CN-FUTURE-MAIN-1d"))
    main_result = _result(
        "FUTURE_MAIN_BAR",
        "国内期货主力K线",
        "akshare",
        "PASS" if not errors and bars else "PARTIAL_FAILURE",
        len(bars),
        f"主力品种{len(selected)}个，日线={len(bars)}；错误={'; '.join(errors) or '无'}",
        "; ".join(errors) or None,
        plan,
    )

    breadth_plan = PersistPlan()
    breadth_rows = 0
    if bars:
        by_day: dict[str, list[dict[str, Any]]] = {}
        for bar in bars:
            by_day.setdefault(str(bar["trading_date"]), []).append(bar)
        snapshots = compute_futures_breadth(by_day, series_kind="MAIN")
        for snapshot in snapshots:
            for key, metric_name in (
                ("ADVANCES", "\u4e0a\u6da8\u5bb6\u6570"),
                ("DECLINES", "\u4e0b\u8dcc\u5bb6\u6570"),
                ("UNCHANGED", "\u5e73\u76d8\u5bb6\u6570"),
            ):
                breadth_plan.gold_metrics.append(
                    _metric(
                        "FUTURES_BREADTH",
                        "CN.FUTURE.BREADTH.MAIN",
                        snapshot.trading_day,
                        "1d",
                        metric_name,
                        getattr(snapshot, key.lower()),
                        snapshot.metric_definition,
                        snapshot.calculation_method,
                        metric_key=key,
                        timestamp=snapshot.timestamp,
                    )
                )
        breadth_rows = len(breadth_plan.gold_metrics)
    breadth_result = _result(
        "FUTURES_BREADTH",
        "\u671f\u8d27\u6bcf\u65e5\u6da8\u8dcc\u5bb6\u6570",
        "local-computed",
        "PASS" if breadth_rows else "BLOCKED",
        breadth_rows,
        f"\u7531{len(selected)}\u4e2a\u4e3b\u529b\u54c1\u79cd\u65e5\u7ebf\u8ba1\u7b97\uff0c\u8986\u76d6{len(set(b['trading_date'] for b in bars))}\u4e2a\u4ea4\u6613\u65e5",
        None if breadth_rows else "no futures bars to compute breadth",
        breadth_plan,
    )
    return [main_result, breadth_result]


def _collect_futures_index() -> CollectionTaskResult:
    api = _ak()
    try:
        frame = api.futures_index_ccidx()
    except Exception as error:
        return _result(
            "FUTURE_INDEX_BAR",
            "\u56fd\u5185\u5546\u54c1\u6307\u6570K\u7ebf",
            "akshare",
            "FAILED",
            0,
            f"{type(error).__name__}: {error}",
            str(error),
            PersistPlan(),
        )
    bars: list[dict[str, Any]] = []
    for row in frame.to_dict(orient="records"):
        code = str(row["\u6307\u6570\u4ee3\u7801"])
        close = _num(row.get("\u6536\u76d8\u70b9\u4f4d")) or _num(row.get("\u7ed3\u7b97\u70b9\u4f4d"))
        bars.append(
            _daily_bar(
                f"CN.CCIDX.INDEX.{code}",
                f"\u5546\u54c1\u6307\u6570{code}",
                row["\u65e5\u671f"],
                open_=row.get("openingPrice"),
                high=row.get("highPrice"),
                low=row.get("lowPrice"),
                close=close,
                volume=None,
                amount=None,
                pct_change=row.get("\u6da8\u8dcc\u5e45"),
                source="akshare-futures_index_ccidx",
            )
        )
    plan = PersistPlan()
    if bars:
        plan.bar_writes.extend(_bar_partitions("CN", "INDEX", "1d", bars, "CN-FUTURE-INDEX-1d"))
    return _result(
        "FUTURE_INDEX_BAR",
        "\u56fd\u5185\u5546\u54c1\u6307\u6570K\u7ebf",
        "akshare",
        "PASS" if bars else "FAILED",
        len(bars),
        f"\u5546\u54c1\u6307\u6570\u65e5\u7ebf={len(bars)}",
        None,
        plan,
    )


def _collect_oi_leaderboard() -> CollectionTaskResult:
    api = _ak()
    day = _last_trading_day_compact()
    trading_day = f"{day[:4]}-{day[4:6]}-{day[6:]}"
    collected_at = _now()
    ranks, coverage = collect_exchange_member_position_ranks(
        api,
        day_compact=day,
        trading_day=trading_day,
        collected_at=collected_at,
    )
    rows = _leaderboard_rows_from_member_ranks(ranks)
    plan = PersistPlan()
    plan.member_position_ranks.extend(rank.to_dict() for rank in ranks)
    plan.member_position_coverage.extend(
        item.to_dict(trading_day=trading_day, collected_at=collected_at) for item in coverage
    )
    if rows:
        leaderboard = build_open_interest_leaderboard(rows, trading_day=trading_day)
        for leader in leaderboard:
            for key, metric_name, attribute_name in (
                ("LONG", "\u591a\u5934\u6301\u4ed3", "long_position"),
                ("LONG_CHANGE", "\u591a\u5934\u6301\u4ed3\u589e\u51cf", "long_position_change"),
                ("SHORT", "\u7a7a\u5934\u6301\u4ed3", "short_position"),
                ("SHORT_CHANGE", "\u7a7a\u5934\u6301\u4ed3\u589e\u51cf", "short_position_change"),
                ("NET", "\u51c0\u6301\u4ed3", "net_position"),
                ("NET_CHANGE", "\u51c0\u6301\u4ed3\u589e\u51cf", "net_position_change"),
            ):
                plan.gold_metrics.append(
                    _metric(
                        "FUTURES_OI_LEADERBOARD",
                        leader.instrument_id,
                        leader.trading_day,
                        "1d",
                        metric_name,
                        getattr(leader, attribute_name),
                        leader.metric_definition,
                        leader.calculation_method,
                        metric_key=key,
                    )
                )
    return _result(
        "FUTURES_OI_LEADERBOARD",
        "\u671f\u8d27\u6301\u4ed3\u9f99\u864e\u699c",
        "akshare",
        "PASS" if ranks and all(item.status == "PASS" for item in coverage)
        else "PARTIAL_FAILURE" if ranks else "FAILED",
        len(ranks),
        f"\u4ea4\u6613\u65e5={trading_day}\uff1b\u516c\u5f00\u6392\u540d={len(ranks)}\uff1b"
        + "; ".join(
            f"{item.exchange}={item.status}({item.contract_count}\u5408\u7ea6/{item.record_count}\u6392\u540d)"
            for item in coverage
        ),
        "; ".join(
            f"{item.exchange}:{item.error or item.status}"
            for item in coverage
            if item.status != "PASS"
        ) or None,
        plan,
    )


def _leaderboard_rows_from_member_ranks(ranks: Sequence[MemberPositionRank]) -> list[MemberPositionRow]:
    """Adapt rank rows to the legacy contract aggregate without inventing member detail."""

    rows: list[MemberPositionRow] = []
    for rank in ranks:
        rows.append(
            MemberPositionRow(
                member=rank.member_name,
                instrument_id=f"CN.{rank.exchange}.FUTURE.{rank.contract_code}",
                trading_day=rank.trading_day,
                long_position=rank.position if rank.side == "LONG" else 0.0,
                long_position_change=(
                    rank.position_change
                    if rank.side == "LONG" and rank.position_change is not None
                    else 0.0
                ),
                short_position=rank.position if rank.side == "SHORT" else 0.0,
                short_position_change=(
                    rank.position_change
                    if rank.side == "SHORT" and rank.position_change is not None
                    else 0.0
                ),
                source=rank.source,
            )
        )
    return rows


def _collect_margin() -> CollectionTaskResult:
    api = _ak()
    today = _last_trading_day_compact()
    errors: list[str] = []
    sh_records: list[dict[str, Any]] = []
    sz_records: list[dict[str, Any]] = []
    bj_records: list[dict[str, Any]] = []
    try:
        sh_records = api.stock_margin_sse(start_date="20260101", end_date=today).to_dict(orient="records")
    except Exception as error:
        errors.append(f"stock_margin_sse:{type(error).__name__}")
    try:
        sz_records = api.macro_china_market_margin_sz().to_dict(orient="records")
    except Exception as error:
        errors.append(f"margin_sz:{type(error).__name__}")
    bj_records: list[dict[str, Any]] = []
    bj_day = ""
    try:
        candidates = [today]
        cursor = date.today()
        while len(candidates) < 5:
            cursor -= timedelta(days=1)
            if cursor.weekday() < 5:
                candidates.append(cursor.isoformat().replace("-", ""))
        for candidate in candidates:
            frame = api.stock_margin_bse(date=candidate)
            bj_records = frame.to_dict(orient="records") if frame is not None else []
            if bj_records:
                bj_day = candidate
                break
    except Exception as error:
        errors.append(f"stock_margin_bse:{type(error).__name__}")
    total_cap_yi: float | None = None
    try:
        spot = api.stock_zh_a_spot_tx().to_dict(orient="records")
        total_cap_yi = sum(_num(row.get("zsz"), 0.0) or 0.0 for row in spot)
    except Exception as error:
        errors.append(f"spot:{type(error).__name__}")

    metrics: list[dict[str, Any]] = []
    sh_day = ""
    if sh_records:
        latest, sh_raw_day = _latest_by_date(sh_records, ("\u4fe1\u7528\u4ea4\u6613\u65e5\u671f",))
        sh_day = f"{sh_raw_day[:4]}-{sh_raw_day[4:6]}-{sh_raw_day[6:]}"
        for key, metric_name in (
            ("\u878d\u8d44\u4f59\u989d", "\u6caa\u5e02\u878d\u8d44\u4f59\u989d"),
            ("\u878d\u8d44\u4e70\u5165\u989d", "\u6caa\u5e02\u878d\u8d44\u4e70\u5165\u989d"),
            ("\u878d\u5238\u4f59\u91cf\u91d1\u989d", "\u6caa\u5e02\u878d\u5238\u4f59\u989d"),
            ("\u878d\u5238\u5356\u51fa\u91cf", "\u6caa\u5e02\u878d\u5238\u5356\u51fa\u91cf"),
            ("\u878d\u8d44\u878d\u5238\u4f59\u989d", "\u6caa\u5e02\u878d\u8d44\u878d\u5238\u4f59\u989d"),
        ):
            value = _num(latest.get(key))
            if value is not None:
                metrics.append(
                    _metric(
                        "CN_MARGIN",
                        "CN.SSE.MARGIN",
                        sh_day,
                        "1d",
                        metric_name,
                        value,
                        "\u4e0a\u4ea4\u6240\u878d\u8d44\u878d\u5238\u660e\u7ec6",
                        "source=stock_margin_sse",
                        metric_key=key,
                    )
                )
    sz_day = ""
    if sz_records:
        latest, sz_day = _latest_by_date(sz_records, ("\u65e5\u671f",))
        for key, metric_name in (
            ("\u878d\u8d44\u4f59\u989d", "\u6df1\u5e02\u878d\u8d44\u4f59\u989d"),
            ("\u878d\u8d44\u4e70\u5165\u989d", "\u6df1\u5e02\u878d\u8d44\u4e70\u5165\u989d"),
            ("\u878d\u5238\u4f59\u989d", "\u6df1\u5e02\u878d\u5238\u4f59\u989d"),
            ("\u878d\u5238\u5356\u51fa\u91cf", "\u6df1\u5e02\u878d\u5238\u5356\u51fa\u91cf"),
            ("\u878d\u8d44\u878d\u5238\u4f59\u989d", "\u6df1\u5e02\u878d\u8d44\u878d\u5238\u4f59\u989d"),
        ):
            value = _num(latest.get(key))
            if value is not None:
                metrics.append(
                    _metric(
                        "CN_MARGIN",
                        "CN.SZSE.MARGIN",
                        sz_day,
                        "1d",
                        metric_name,
                        value,
                        "\u6df1\u4ea4\u6240\u878d\u8d44\u878d\u5238\u660e\u7ec6",
                        "source=macro_china_market_margin_sz",
                        metric_key=key,
                    )
                )
    bj_metrics: list[tuple[str, str]] = [
        ("\u878d\u8d44\u4f59\u989d", "\u4eac\u5e02\u878d\u8d44\u4f59\u989d"),
        ("\u878d\u8d44\u4e70\u5165\u989d", "\u4eac\u5e02\u878d\u8d44\u4e70\u5165\u989d"),
        ("\u878d\u5238\u4f59\u989d", "\u4eac\u5e02\u878d\u5238\u4f59\u989d"),
        ("\u878d\u5238\u4f59\u91cf", "\u4eac\u5e02\u878d\u5238\u4f59\u91cf"),
        ("\u878d\u5238\u5356\u51fa\u91cf", "\u4eac\u5e02\u878d\u5238\u5356\u51fa\u91cf"),
        ("\u878d\u8d44\u878d\u5238\u4f59\u989d", "\u4eac\u5e02\u878d\u8d44\u878d\u5238\u4f59\u989d"),
    ]
    if bj_records:
        latest = bj_records[0]
        bj_date = f"{bj_day[:4]}-{bj_day[4:6]}-{bj_day[6:]}" if bj_day else ""
        for key, metric_name in bj_metrics:
            value = _num(latest.get(key))
            if value is not None and bj_date:
                metrics.append(
                    _metric(
                        "CN_MARGIN",
                        "CN.BSE.MARGIN",
                        bj_date,
                        "1d",
                        metric_name,
                        value,
                        "\u5317\u4ea4\u6240\u878d\u8d44\u878d\u5238\u660e\u7ec6",
                        f"source=stock_margin_bse; date={bj_day}",
                        metric_key=key,
                    )
                )
    if not bj_records:
        errors.append("\u4eac\u5e02\uff1a\u5f53\u65e5\u65e0\u878d\u8d44\u878d\u5238\u6570\u636e\uff08\u7a7a\u8868\uff09")

    sh_latest, _ = _latest_by_date(sh_records, ("\u4fe1\u7528\u4ea4\u6613\u65e5\u671f",))
    sz_latest, _ = _latest_by_date(sz_records, ("\u65e5\u671f",))
    sh_balance = _num(sh_latest.get("\u878d\u8d44\u878d\u5238\u4f59\u989d")) if sh_latest else None
    sz_balance = _num(sz_latest.get("\u878d\u8d44\u878d\u5238\u4f59\u989d")) if sz_latest else None
    ratio_day = sh_day or sz_day
    if sh_balance is not None and sz_balance is not None and total_cap_yi and ratio_day:
        metrics.append(
            _metric(
                "CN_MARGIN",
                "CN.MARGIN.RATIO",
                ratio_day,
                "1d",
                "\u6caa\u6df1\u4e24\u878d\u4f59\u989d\u5360A\u80a1\u603b\u5e02\u503c\u6bd4(\u7ea6)",
                (sh_balance + sz_balance) / (total_cap_yi * 1e8),
                "\u4e24\u878d\u4f59\u989d\u5408\u8ba1 / A\u80a1\u603b\u5e02\u503c(\u817e\u8baf\u5feb\u7167 zsz \u5408\u8ba1)",
                "\u8fd1\u4f3c\u4f30\u7b97\uff1b\u5e02\u503c\u53d6\u603b\u5e02\u503c\u800c\u975e\u6d41\u901a\u5e02\u503c",
                metric_key="MARGIN_CAP_RATIO",
            )
        )
    plan = PersistPlan(gold_metrics=metrics)
    return _result(
        "CN_MARGIN",
        "\u6caa\u6df1\u4eac\u878d\u8d44\u878d\u5238",
        "akshare",
        "PASS" if metrics and not errors else "PARTIAL_FAILURE",
        len(metrics),
        f"\u6caa\u5e02={sh_day or '无'} \u6df1\u5e02={sz_day or '无'} \u4eac\u5e02={len(bj_records)}\u884c\uff1b\u9519\u8bef={'; '.join(errors) or '无'}",
        "; ".join(errors) or None,
        plan,
    )


def _collect_macro() -> CollectionTaskResult:
    api = _ak()
    points: list[MacroPoint] = []
    errors: list[str] = []

    def _fetch_rows(fetcher: Callable[[], Any], label: str) -> list[dict[str, Any]]:
        last_error: Exception | None = None
        for attempt in range(2):
            try:
                return fetcher().to_dict(orient="records")
            except Exception as error:  # noqa: BLE001 - 采集层需要尽量兜底
                last_error = error
                if attempt == 0:
                    time.sleep(2)
        errors.append(f"{label}:{type(last_error).__name__}")
        return []

    try:
        money = api.macro_china_money_supply().to_dict(orient="records")
        for row in money:
            month = _parse_cn_month(row["\u6708\u4efd"])
            m0 = _num(row.get("\u6d41\u901a\u4e2d\u7684\u73b0\u91d1(M0)-\u540c\u6bd4\u589e\u957f"))
            m1 = _num(row.get("\u8d27\u5e01(M1)-\u540c\u6bd4\u589e\u957f"))
            m2 = _num(row.get("\u8d27\u5e01\u548c\u51c6\u8d27\u5e01(M2)-\u540c\u6bd4\u589e\u957f"))
            if m0 is not None:
                points.append(normalise_macro_point("M0_MONEY_SUPPLY", available_time=month, value=m0))
            if m1 is not None:
                points.append(normalise_macro_point("M1_MONEY_SUPPLY", available_time=month, value=m1))
            if m2 is not None:
                points.append(normalise_macro_point("M2_MONEY_SUPPLY", available_time=month, value=m2))
    except Exception as error:
        errors.append(f"money_supply:{type(error).__name__}")

    for series_id, label, fetcher in (
        ("CN_IMPORT_USD_YOY", "cn_import_usd_yoy", lambda: api.macro_china_imports_yoy()),
        ("CN_EXPORT_USD_YOY", "cn_export_usd_yoy", lambda: api.macro_china_exports_yoy()),
        ("CN_TRADE_BALANCE_USD", "cn_trade_balance_usd", lambda: api.macro_china_trade_balance()),
        ("CN_FOREX_RESERVES", "cn_forex_reserves", lambda: api.macro_china_fx_reserves_yearly()),
    ):
        for row in _fetch_rows(fetcher, label):
            value = _num(row.get("今值"))
            if value is not None and row.get("日期") is not None:
                points.append(normalise_macro_point(series_id, available_time=row["日期"], value=value))

    for row in _fetch_rows(
        lambda: api.macro_china_consumer_goods_retail(),
        "cn_consumer_goods_retail",
    ):
        month = _parse_cn_month(row.get("月份"))
        yoy = _num(row.get("同比增长"))
        mom = _num(row.get("环比增长"))
        if yoy is not None:
            points.append(normalise_macro_point("CN_RETAIL_SALES_YOY", available_time=month, value=yoy))
        if mom is not None:
            points.append(normalise_macro_point("CN_RETAIL_SALES_MOM", available_time=month, value=mom))

    for row in _fetch_rows(api.macro_china_society_electricity, "cn_society_electricity"):
        month = _parse_cn_month(row.get("统计时间"))
        raw_total = _num(row.get("全社会用电量"))
        if raw_total is not None and month:
            # The local AkShare adapter relays Sina's table in 10,000 kWh.
            # NEA's 2026-07 bulletin gives the independent scale check:
            # 613,990,000 / 10,000 = 61,399 亿千瓦时 (January–July cumulative).
            points.append(
                normalise_macro_point(
                    "CN_ELECTRICITY_CONSUMPTION",
                    available_time=month,
                    value=raw_total / 10_000.0,
                )
            )

    try:
        for row in api.macro_china_cpi_yearly().to_dict(orient="records"):
            value = _num(row.get("\u4eca\u503c"))
            if value is not None:
                points.append(normalise_macro_point("CPI", available_time=row["\u65e5\u671f"], value=value))
    except Exception as error:
        errors.append(f"cpi:{type(error).__name__}")
    try:
        for row in api.macro_china_ppi_yearly().to_dict(orient="records"):
            value = _num(row.get("\u4eca\u503c"))
            if value is not None:
                points.append(normalise_macro_point("PPI", available_time=row["\u65e5\u671f"], value=value))
    except Exception as error:
        errors.append(f"ppi:{type(error).__name__}")

    try:
        for row in api.macro_china_pmi().to_dict(orient="records"):
            month = _parse_cn_month(row["\u6708\u4efd"])
            manufacturing = _num(row.get("\u5236\u9020\u4e1a-\u6307\u6570"))
            services = _num(row.get("\u975e\u5236\u9020\u4e1a-\u6307\u6570"))
            if manufacturing is not None:
                points.append(normalise_macro_point("PMI_MANUFACTURING", available_time=month, value=manufacturing))
            if services is not None:
                points.append(normalise_macro_point("PMI_SERVICES", available_time=month, value=services))
    except Exception as error:
        errors.append(f"pmi:{type(error).__name__}")
    for row in _fetch_rows(api.macro_china_cx_pmi_yearly, "cx_pmi"):
        value = _num(row.get("\u4eca\u503c"))
        if value is not None:
            points.append(
                normalise_macro_point("PMI_CAIXIN_MANUFACTURING", available_time=row["\u65e5\u671f"], value=value)
            )
    for row in _fetch_rows(api.macro_china_cx_services_pmi_yearly, "cx_services"):
        value = _num(row.get("\u4eca\u503c"))
        if value is not None:
            points.append(
                normalise_macro_point("PMI_CAIXIN_SERVICES", available_time=row["\u65e5\u671f"], value=value)
            )

    try:
        for row in api.repo_rate_query(
            symbol="\u94f6\u94f6\u95f4\u56de\u8d2d\u5b9a\u76d8\u5229\u7387"
        ).to_dict(orient="records"):
            value = _num(row.get("FDR007")) or _num(row.get("FR007"))
            if value is not None:
                points.append(normalise_macro_point("DR007", available_time=row["date"], value=value))
    except Exception as error:
        errors.append(f"repo_rate:{type(error).__name__}")

    try:
        for row in api.bond_zh_us_rate().to_dict(orient="records"):
            cn = _num(row.get("\u4e2d\u56fd\u56fd\u503a\u6536\u76ca\u738710\u5e74"))
            us = _num(row.get("\u7f8e\u56fd\u56fd\u503a\u6536\u76ca\u738710\u5e74"))
            if cn is not None:
                points.append(normalise_macro_point("CN10Y_YIELD", available_time=row["\u65e5\u671f"], value=cn))
            if us is not None:
                points.append(normalise_macro_point("US10Y_YIELD", available_time=row["\u65e5\u671f"], value=us))
    except Exception as error:
        errors.append(f"bond_rate:{type(error).__name__}")

    try:
        for row in api.macro_bank_usa_interest_rate().to_dict(orient="records"):
            value = _num(row.get("\u4eca\u503c"))
            if value is not None:
                points.append(normalise_macro_point("FED_FUNDS_RATE", available_time=row["\u65e5\u671f"], value=value))
    except Exception as error:
        errors.append(f"fed_rate:{type(error).__name__}")

    for row in _fetch_rows(lambda: api.macro_usa_non_farm(), "us_nonfarm_payrolls"):
        value = _num(row.get("今值"))
        if value is not None and row.get("日期") is not None:
            points.append(
                normalise_macro_point(
                    "US_NONFARM_PAYROLLS_SA",
                    available_time=row["日期"],
                    value=value * 10.0,
                )
            )

    try:
        for month, value in _fetch_bea_us_imports_rows():
            points.append(
                normalise_macro_point(
                    "US_IMPORTS_SA",
                    available_time=month,
                    value=value,
                    source="美国经济分析局 / 人口普查局",
                )
            )
    except Exception as error:  # noqa: BLE001 - remote release availability is non-deterministic
        errors.append(f"us_imports_sa:{type(error).__name__}")

    try:
        for quarter, value in _fetch_bea_core_pce_final_rows():
            points.append(
                normalise_macro_point(
                    "US_CORE_PCE_QOQ_SAAR_FINAL",
                    available_time=quarter,
                    value=value,
                    source="美国经济分析局",
                )
            )
    except Exception as error:  # noqa: BLE001 - remote release availability is non-deterministic
        errors.append(f"us_core_pce_qoq_saar_final:{type(error).__name__}")

    cpi = [point for point in points if point.series_id == "CPI"]
    ppi = [point for point in points if point.series_id == "PPI"]
    if cpi and ppi:
        points.extend(
            derive_series(
                {"CPI": cpi, "PPI": ppi},
                derived_series_id="CPI_PPI_SPREAD",
                formula="A-B",
            )
        )
    metrics = _macro_metrics(points)
    plan = PersistPlan(gold_metrics=metrics)
    return _result(
        "MACRO_SERIES",
        "\u5b8f\u89c2\u6570\u636e\u5e8f\u5217",
        "akshare",
        "PASS" if metrics and not errors else "PARTIAL_FAILURE",
        len(metrics),
        f"M0/M1/M2/CPI/PPI/PMI/DR007/\u4e2d\u7f8e10Y/\u7f8e\u8054\u50a8\u5229\u7387\uff0c\u5171{len(metrics)}\u6761\uff1b\u9519\u8bef={'; '.join(errors) or '无'}",
        "; ".join(errors) or None,
        plan,
    )


def _collect_a_share_breadth() -> CollectionTaskResult:
    api = _ak()
    try:
        records = api.stock_zh_a_spot_tx().to_dict(orient="records")
    except Exception as error:
        return _result(
            "A_SHARE_BREADTH",
            "A\u80a1\u6da8\u8dcc/\u6da8\u505c/\u5e02\u503c\u5feb\u7167",
            "akshare",
            "FAILED",
            0,
            f"{type(error).__name__}: {error}",
            str(error),
            PersistPlan(),
        )
    changes = [_num(row.get("zdf")) for row in records if _num(row.get("zdf")) is not None]
    advances = sum(1 for change in changes if change > 0)
    declines = sum(1 for change in changes if change < 0)
    unchanged = len(changes) - advances - declines
    total_cap_yi = sum(_num(row.get("zsz"), 0.0) or 0.0 for row in records)
    total_amount_yi = sum(_num(row.get("turnover"), 0.0) or 0.0 for row in records) / 10000.0
    day = _last_trading_day_compact()
    day = f"{day[:4]}-{day[4:6]}-{day[6:]}"
    plan = PersistPlan(
        gold_metrics=[
            _metric(
                "A_SHARE_BREADTH",
                "CN.A_SHARE.BREADTH",
                day,
                "1d",
                "\u4e0a\u6da8\u5bb6\u6570",
                advances,
                "\u817e\u8baf\u5168\u5e02\u573a\u5feb\u7167\u6da8\u8dcc\u5e45\u7edf\u8ba1",
                "source=stock_zh_a_spot_tx; close>prev=up",
                metric_key="ADVANCES",
            ),
            _metric(
                "A_SHARE_BREADTH",
                "CN.A_SHARE.BREADTH",
                day,
                "1d",
                "\u4e0b\u8dcc\u5bb6\u6570",
                declines,
                "\u817e\u8baf\u5168\u5e02\u573a\u5feb\u7167\u6da8\u8dcc\u5e45\u7edf\u8ba1",
                "source=stock_zh_a_spot_tx; close<prev=down",
                metric_key="DECLINES",
            ),
            _metric(
                "A_SHARE_BREADTH",
                "CN.A_SHARE.BREADTH",
                day,
                "1d",
                "\u5e73\u76d8\u5bb6\u6570",
                unchanged,
                "\u817e\u8baf\u5168\u5e02\u573a\u5feb\u7167\u6da8\u8dcc\u5e45\u7edf\u8ba1",
                "source=stock_zh_a_spot_tx",
                metric_key="UNCHANGED",
            ),
            _metric(
                "A_SHARE_BREADTH",
                "CN.A_SHARE.BREADTH",
                day,
                "1d",
                "\u6caa\u6df1\u4eac\u603b\u5e02\u503c(\u4ebf)",
                total_cap_yi,
                "\u5168\u5e02\u573a zsz \u603b\u5e02\u503c\u5408\u8ba1",
                "source=stock_zh_a_spot_tx; unit=\u4ebf\u5143",
                metric_key="TOTAL_MARKET_CAP_YI",
            ),
            _metric(
                "A_SHARE_BREADTH",
                "CN.A_SHARE.BREADTH",
                day,
                "1d",
                "\u5f53\u65e5\u6210\u4ea4\u989d(\u4ebf)",
                total_amount_yi,
                "\u5168\u5e02\u573a turnover \u5408\u8ba1",
                "source=stock_zh_a_spot_tx; unit=\u4ebf\u5143",
                metric_key="TOTAL_AMOUNT_YI",
            ),
        ]
    )
    return _result(
        "A_SHARE_BREADTH",
        "A\u80a1\u6da8\u8dcc/\u5e02\u503c\u5feb\u7167",
        "akshare",
        "PASS" if records else "FAILED",
        len(plan.gold_metrics),
        f"\u5168\u5e02\u573a{len(records)}\u53ea\uff1b\u6da8={advances} \u8dcc={declines} \u5e73={unchanged}",
        None,
        plan,
    )


def _collect_zt_pool() -> CollectionTaskResult:
    api = _ak()
    day = _last_trading_day_compact()
    metrics: list[dict[str, Any]] = []
    errors: list[str] = []
    for function_name, metric_prefix, level_field in (
        ("stock_zt_pool_em", "\u6da8\u505c", "\u8fde\u677f\u6570"),
        ("stock_zt_pool_dtgc_em", "\u8dcc\u505c", "\u8fde\u7eed\u8dcc\u505c"),
    ):
        try:
            frame = getattr(api, function_name)(date=day)
        except Exception as error:
            errors.append(f"{function_name}:{type(error).__name__}")
            continue
        records = frame.to_dict(orient="records")
        metrics.append(
            _metric(
                "CN_ZT_POOL",
                f"CN.A_SHARE.{metric_prefix}",
                day,
                "1d",
                f"{metric_prefix}\u5bb6\u6570",
                len(records),
                "\u4e1c\u8d22\u6da8\u505c/\u8dcc\u505c\u6c60",
                f"source={function_name}",
                metric_key=f"{metric_prefix}_COUNT",
            )
        )
        if records:
            level_values = [
                _num(row.get(level_field)) or _num(row.get("\u8fde\u677f\u6570")) or 0.0
                for row in records
            ]
            max_level = max(level_values)
            metrics.append(
                _metric(
                    "CN_ZT_POOL",
                    f"CN.A_SHARE.{metric_prefix}",
                    day,
                    "1d",
                    f"{metric_prefix}\u6700\u9ad8\u8fde\u677f\u9ad8\u5ea6",
                    max_level,
                    "\u8fde\u677f\u6570\u6700\u5927\u503c",
                    f"source={function_name}",
                    metric_key=f"{metric_prefix}_MAX_LIANBAN",
                )
            )
    prev_records: list[dict[str, Any]] = []
    try:
        prev_records = api.stock_zt_pool_previous_em(date=day).to_dict(orient="records")
    except Exception as error:
        errors.append(f"previous_pool:{type(error).__name__}")
    if prev_records:
        specs: list[tuple[str, int, str]] = []
        for row in prev_records:
            code = str(row.get("\u4ee3\u7801", "")).zfill(6)
            market = 1 if code.startswith(("6", "9")) else 2 if code.startswith(("4", "8")) else 0
            specs.append((code, market, str(row.get("\u540d\u79f0", ""))))
        bars, _bar_errors = _tdx_bars("1d", specs, 2, asset_type="STOCK")
        target_day = f"{day[:4]}-{day[4:6]}-{day[6:]}"
        bars_by_code: dict[str, dict[str, Any]] = {}
        for bar in bars:
            if str(bar.get("trading_date", "")) == target_day:
                bars_by_code.setdefault(str(bar.get("symbol", "")), bar)
        returns: list[float] = []
        for code, _, _ in specs:
            bar = bars_by_code.get(code)
            if not bar:
                continue
            open_price = _num(bar.get("open"))
            close = _num(bar.get("close"))
            if open_price and close:
                returns.append((close - open_price) / open_price * 100.0)
        return_value: float | None = None
        return_detail = ""
        if len(returns) >= max(3, len(specs) // 2):
            return_value = sum(returns) / len(returns)
            return_detail = (
                f"pytdx 昨日涨停池{len(specs)}只中{len(returns)}只有当日K线，按(收盘-开盘)/开盘取均值；单位%"
            )
        else:
            changes = [
                _num(row.get("\u6da8\u8dcc\u5e45"))
                for row in prev_records
                if _num(row.get("\u6da8\u8dcc\u5e45")) is not None
            ]
            if changes:
                return_value = sum(changes) / len(changes)
                return_detail = (
                    f"K线覆盖不足({len(returns)}/{len(specs)})，回退用东财昨日涨停池涨跌幅(今收/昨收-1)均值；单位%"
                )
        if return_value is not None:
            metrics.append(
                _metric(
                    "CN_ZT_POOL",
                    "CN.A_SHARE.YESTERDAY_LIMIT_UP",
                    day,
                    "1d",
                    "昨日涨停今日接盘收益率(%)",
                    return_value,
                    "昨日涨停个股今日开盘价买入的当日平均收益率",
                    return_detail,
                    metric_key="YESTERDAY_LIMIT_UP_OPEN_RETURN",
                )
            )
    plan = PersistPlan(gold_metrics=metrics)
    return _result(
        "CN_ZT_POOL",
        "\u6da8\u505c/\u8dcc\u505c\u6c60\u4e0e\u8fde\u677f\u9ad8\u5ea6",
        "akshare",
        "PASS" if metrics and not errors else "PARTIAL_FAILURE",
        len(metrics),
        f"{day} \u6da8\u505c/\u8dcc\u505c\u6c60\uff1b\u9519\u8bef={'; '.join(errors) or '无'}",
        "; ".join(errors) or None,
        plan,
    )


def _collect_hsgt() -> CollectionTaskResult:
    api = _ak()
    metrics: list[dict[str, Any]] = []
    errors: list[str] = []
    try:
        summary = api.stock_hsgt_fund_flow_summary_em().to_dict(orient="records")
        for row in summary:
            day = str(row["\u4ea4\u6613\u65e5"])
            board = str(row["\u677f\u5757"])
            direction = str(row["\u8d44\u91d1\u65b9\u5411"])
            instrument_id = f"CN.HSGT.{board}.{direction}"
            for key, metric_name in (
                ("\u6210\u4ea4\u51c0\u4e70\u989d", "\u6210\u4ea4\u51c0\u4e70\u989d(\u4ebf)"),
                ("\u8d44\u91d1\u51c0\u6d41\u5165", "\u8d44\u91d1\u51c0\u6d41\u5165(\u4ebf)"),
                ("\u4e0a\u6da8\u6570", "\u4e0a\u6da8\u5bb6\u6570"),
                ("\u4e0b\u8dcc\u6570", "\u4e0b\u8dcc\u5bb6\u6570"),
            ):
                value = _num(row.get(key))
                if value is not None:
                    metrics.append(
                        _metric(
                            "HSGT_FLOW",
                            instrument_id,
                            day,
                            "1d",
                            metric_name,
                            value,
                            "\u6e2f\u80a1\u901a/\u6df1\u80a1\u901a\u8d44\u91d1\u6d41\u52a8",
                            "source=stock_hsgt_fund_flow_summary_em",
                            metric_key=f"{board}_{direction}_{key}",
                        )
                    )
    except Exception as error:
        errors.append(f"summary:{type(error).__name__}")
    for direction, symbol in (("\u5317\u5411", "\u5317\u5411\u8d44\u91d1"), ("\u5357\u5411", "\u5357\u5411\u8d44\u91d1")):
        try:
            history = api.stock_hsgt_hist_em(symbol=symbol).to_dict(orient="records")
        except Exception as error:
            errors.append(f"hist_{direction}:{type(error).__name__}")
            continue
        if not history:
            continue
        latest = history[-1]
        day = str(latest["\u65e5\u671f"])
        for key, metric_name in (
            ("\u5f53\u65e5\u6210\u4ea4\u51c0\u4e70\u989d", f"{direction}\u5f53\u65e5\u6210\u4ea4\u51c0\u4e70\u989d(\u4ebf)"),
            ("\u5386\u53f2\u7d2f\u8ba1\u51c0\u4e70\u989d", f"{direction}\u5386\u53f2\u7d2f\u8ba1\u51c0\u4e70\u989d(\u4ebf)"),
            ("\u6301\u80a1\u5e02\u503c", f"{direction}\u6301\u80a1\u5e02\u503c(\u4ebf)"),
        ):
            value = _num(latest.get(key))
            if value is not None:
                metrics.append(
                    _metric(
                        "HSGT_FLOW",
                        f"CN.HSGT.{direction}",
                        day,
                        "1d",
                        metric_name,
                        value,
                        "\u6e2f\u80a1\u901a/\u6df1\u80a1\u901a\u5386\u53f2\u8d44\u91d1\u6d41\u52a8",
                        "source=stock_hsgt_hist_em",
                        metric_key=f"{direction}_{key}",
                    )
                )
    plan = PersistPlan(gold_metrics=metrics)
    return _result(
        "HSGT_FLOW",
        "\u5317\u5411\u5357\u5411\u8d44\u91d1",
        "akshare",
        "PASS" if metrics and not errors else "PARTIAL_FAILURE",
        len(metrics),
        f"\u5317\u5411/\u5357\u5411\u8d44\u91d1\u6307\u6807={len(metrics)}\uff1b\u9519\u8bef={'; '.join(errors) or '无'}",
        "; ".join(errors) or None,
        plan,
    )


def _collect_foreign_futures() -> CollectionTaskResult:
    api = _ak()
    specs = [
        ("GC", "COMEX\u9ec4\u91d1", "US.COMEX.FUTURE.GC"),
        ("SI", "COMEX\u767d\u94f6", "US.COMEX.FUTURE.SI"),
        ("HG", "COMEX\u94dc", "US.COMEX.FUTURE.HG"),
        ("CL", "WTI\u539f\u6cb9", "US.NYMEX.FUTURE.CL"),
    ]
    bars: list[dict[str, Any]] = []
    errors: list[str] = []
    closes: dict[str, list[MacroPoint]] = {}
    for symbol, name, instrument_id in specs:
        try:
            frame = api.futures_foreign_hist(symbol=symbol)
        except Exception as error:
            errors.append(f"{symbol}:{type(error).__name__}")
            continue
        records = frame.to_dict(orient="records")
        for row in records:
            day = str(row["date"])[:10]
            bars.append(
                _daily_bar(
                    instrument_id,
                    name,
                    day,
                    open_=row.get("open"),
                    high=row.get("high"),
                    low=row.get("low"),
                    close=row.get("close"),
                    volume=row.get("volume"),
                    amount=None,
                    open_interest=row.get("position"),
                    currency="USD",
                    source="akshare-futures_foreign_hist",
                )
            )
        closes[symbol] = [
            MacroPoint(
                series_id=symbol,
                name=name,
                frequency="DAILY",
                unit="USD",
                source="akshare-futures_foreign_hist",
                available_time=str(row["date"])[:10],
                value=float(row["close"]),
                definition="\u5916\u76d8\u671f\u8d27\u6536\u76d8\u4ef7",
                calculation_method="futures_foreign_hist close",
                fetched_at=_now(),
            )
            for row in records
        ]
    metrics: list[dict[str, Any]] = []
    try:
        realtime = api.futures_foreign_commodity_realtime(symbol=["GC", "SI", "HG"]).to_dict(orient="records")
        name_to_symbol = {"COMEX\u9ec4\u91d1": "GC", "COMEX\u767d\u94f6": "SI", "COMEX\u94dc": "HG"}
        symbol_to_instrument = {symbol: instrument_id for symbol, _, instrument_id in specs}
        for row in realtime:
            symbol = name_to_symbol.get(str(row["\u540d\u79f0"]))
            if symbol is None:
                continue
            instrument_id = f"{symbol_to_instrument[symbol]}.REALTIME"
            day = str(row.get("\u65e5\u671f") or "")[:10]
            for key, metric_name in (
                ("\u6700\u65b0\u4ef7", "\u6700\u65b0\u4ef7"),
                ("\u6da8\u8dcc\u5e45", "\u6da8\u8dcc\u5e45(%)"),
                ("\u6301\u4ed3\u91cf", "\u6301\u4ed3\u91cf"),
            ):
                value = _num(row.get(key))
                if value is not None:
                    metrics.append(
                        _metric(
                            "FUTURE_GLOBAL_BAR",
                            instrument_id,
                            day,
                            "1d",
                            metric_name,
                            value,
                            "\u5916\u76d8\u671f\u8d27\u5b9e\u65f6\u5feb\u7167",
                            "source=futures_foreign_commodity_realtime",
                            metric_key=f"{symbol}_{key}",
                        )
                    )
    except Exception as error:
        errors.append(f"realtime:{type(error).__name__}")
    if "GC" in closes and "SI" in closes:
        metrics.extend(
            _macro_metrics(
                derive_series(
                    {"GOLD": closes["GC"], "SILVER": closes["SI"]},
                    derived_series_id="GOLD_SILVER_RATIO",
                    formula="A/B",
                )
            )
        )
    if "GC" in closes and "CL" in closes:
        metrics.extend(
            _macro_metrics(
                derive_series(
                    {"GOLD": closes["GC"], "OIL": closes["CL"]},
                    derived_series_id="GOLD_OIL_RATIO",
                    formula="A/B",
                )
            )
        )
    plan = PersistPlan()
    if bars:
        plan.bar_writes.extend(_bar_partitions("GLOBAL", "FUTURE", "1d", bars, "GLOBAL-FUTURE-1d"))
    plan.gold_metrics.extend(metrics)
    return _result(
        "FUTURE_GLOBAL_BAR",
        "\u5916\u76d8\u671f\u8d27/\u91d1\u94f6\u6bd4/\u91d1\u6cb9\u6bd4",
        "akshare",
        "PASS" if bars and not errors else "PARTIAL_FAILURE",
        len(bars) + len(metrics),
        f"\u5916\u76d8K\u7ebf={len(bars)} \u6307\u6807={len(metrics)}\uff1b\u9519\u8bef={'; '.join(errors) or '无'}",
        "; ".join(errors) or None,
        plan,
    )


def _collect_crypto() -> CollectionTaskResult:
    metrics: list[dict[str, Any]] = []
    bars: list[dict[str, Any]] = []
    errors: list[str] = []
    for symbol, name, series_id in (("BTCUSDT", "\u6bd4\u7279\u5e01", "BTC_USD"), ("ETHUSDT", "\u4ee5\u592a\u574a", "ETH_USD")):
        url = f"https://data-api.binance.vision/api/v3/klines?symbol={symbol}&interval=1d&limit=60"
        try:
            request = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with _direct_network():
                with urllib.request.urlopen(request, timeout=20) as response:
                    klines = json.loads(response.read().decode("utf-8"))
        except Exception as error:
            errors.append(f"{symbol}:{type(error).__name__}")
            continue
        for kline in klines:
            day = datetime.fromtimestamp(int(kline[0]) / 1000, tz=timezone.utc).strftime("%Y-%m-%d")
            bars.append(
                _daily_bar(
                    f"GLOBAL.CRYPTO.{symbol}",
                    name,
                    day,
                    open_=float(kline[1]),
                    high=float(kline[2]),
                    low=float(kline[3]),
                    close=float(kline[4]),
                    volume=float(kline[5]),
                    amount=float(kline[7]),
                    currency="USD",
                    source="binance-klines",
                )
            )
            metrics.append(
                _metric(
                    series_id,
                    series_id,
                    day,
                    "DAILY",
                    macro_series_index()[series_id].name,
                    float(kline[4]),
                    macro_series_index()[series_id].definition,
                    macro_series_index()[series_id].calculation_method,
                    metric_key="close",
                )
            )
    plan = PersistPlan()
    if bars:
        plan.bar_writes.extend(_bar_partitions("GLOBAL", "CRYPTO", "1d", bars, "GLOBAL-CRYPTO-1d"))
    plan.gold_metrics.extend(metrics)
    return _result(
        "CRYPTO_BAR",
        "\u52a0\u5bc6\u8d27\u5e01K\u7ebf",
        "binance",
        "PASS" if bars and not errors else "PARTIAL_FAILURE",
        len(bars) + len(metrics),
        f"BTC/ETH\u65e5\u7ebf={len(bars)} \u6307\u6807={len(metrics)}\uff1b\u9519\u8bef={'; '.join(errors) or '无'}",
        "; ".join(errors) or None,
        plan,
    )


def _fetch_url_text(url: str, timeout: int = 20) -> str:
    """GET 远端文本：临时绕过系统代理，并对个别证书链不完整的源放宽 TLS 校验。"""

    import requests
    import urllib3

    with _direct_network():
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        response = requests.get(
            url,
            headers={
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                    "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
                ),
                "Accept": "application/json,text/plain,*/*",
                "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
            },
            timeout=timeout,
            verify=False,
        )
        response.raise_for_status()
        return response.text


def _parse_cboe_vix(text: str) -> list[tuple[str, float]]:
    """解析 CBOE VIX 官方 CSV（DATE,OPEN,HIGH,LOW,CLOSE，日期 MM/DD/YYYY）。"""

    rows: list[tuple[str, float]] = []
    for line in text.splitlines():
        if not line.strip() or line.strip().upper().startswith("DATE"):
            continue
        parts = line.split(",")
        if len(parts) < 5:
            continue
        try:
            day = datetime.strptime(parts[0].strip(), "%m/%d/%Y").strftime("%Y-%m-%d")
            close = float(parts[4])
        except ValueError:
            continue
        rows.append((day, close))
    return rows


def _parse_tencent_vix(text: str) -> list[tuple[str, float]]:
    """解析腾讯行情 usVIX 日线：[日期, 开, 收, 高, 低, ...]。"""

    document = json.loads(text)
    rows: list[tuple[str, float]] = []
    for row in document["data"]["usVIX"]["day"]:
        try:
            rows.append((str(row[0]), float(row[2])))
        except (IndexError, TypeError, ValueError):
            continue
    return rows


def _fetch_vix_points() -> list[tuple[str, float]]:
    """VIX 收盘序列：CBOE 官方 CSV 优先，腾讯行情兜底。"""

    last_error: Exception | None = None
    for url, parser in (
        (
            "https://cdn.cboe.com/api/global/us_indices/daily_prices/VIX_History.csv",
            _parse_cboe_vix,
        ),
        (
            "https://web.ifzq.gtimg.cn/appstock/app/usfqkline/get?param=usVIX,day,,,250,qfq",
            _parse_tencent_vix,
        ),
    ):
        try:
            rows = parser(_fetch_url_text(url))
            if rows:
                return rows
            last_error = ValueError("empty VIX rows")
        except Exception as error:  # noqa: BLE001 - 采集层需要尽量兜底
            last_error = error
    raise last_error or ValueError("no VIX source")


def _parse_eastmoney_kline(document: dict[str, Any]) -> list[tuple[str, float]]:
    """解析东方财富日 K 线（f51 日期, f53 收盘）。"""

    data = document.get("data")
    if not data or not data.get("klines"):
        raise ValueError("eastmoney USD index: empty data")
    rows: list[tuple[str, float]] = []
    for line in data["klines"]:
        parts = line.split(",")
        try:
            rows.append((parts[0], float(parts[2])))
        except (IndexError, ValueError):
            continue
    return rows


def _parse_tencent_kline(document: dict[str, Any], symbol: str) -> list[tuple[str, float]]:
    """解析腾讯日 K 线（[日期, 开, 收, 高, 低, 量]）。"""

    data = (document.get("data") or {}).get(symbol) or {}
    rows: list[tuple[str, float]] = []
    for item in data.get("day") or []:
        try:
            rows.append((str(item[0]), float(item[2])))
        except (IndexError, ValueError):
            continue
    if not rows:
        raise ValueError(f"tencent {symbol}: empty klines")
    return rows


def _fetch_usd_index_points() -> list[tuple[str, float]]:
    """美元指数 DXY 序列：东方财富日 K 线优先，腾讯 whDINIW 日 K 线，实时行情兜底。"""

    kline_url = (
        "https://push2his.eastmoney.com/api/qt/stock/kline/get"
        "?secid=100.UDI&klt=101&fqt=1&lmt=250&end=20500101&iscca=1"
        "&fields1=f1,f2,f3,f4,f5,f6"
        "&fields2=f51,f52,f53,f54,f55,f56,f57,f58,f59,f60,f61"
    )
    last_error: Exception | None = None
    for attempt in range(2):
        try:
            rows = _parse_eastmoney_kline(json.loads(_fetch_url_text(kline_url)))
            if rows:
                return rows
            last_error = ValueError("eastmoney USD index: empty klines")
        except Exception as error:  # noqa: BLE001 - 采集层需要尽量兜底
            last_error = error
            if attempt == 0:
                time.sleep(2)
    try:
        return _parse_tencent_kline(
            json.loads(
                _fetch_url_text(
                    "https://web.ifzq.gtimg.cn/appstock/app/fqkline/get"
                    "?param=whDINIW,day,,,320,qfq"
                )
            ),
            "whDINIW",
        )
    except Exception as error:  # noqa: BLE001 - 采集层需要尽量兜底
        last_error = error
    spot_errors: list[str] = []
    for spot_host in ("push2delay.eastmoney.com", "push2.eastmoney.com"):
        try:
            spot = json.loads(
                _fetch_url_text(
                    f"https://{spot_host}/api/qt/stock/get"
                    "?secid=100.UDI&fields=f43,f44,f45,f46,f57,f58,f60,f86,f169,f170"
                )
            )
            data = spot.get("data") or {}
            close = data.get("f46")
            timestamp = data.get("f86")
            if close is None:
                raise ValueError("eastmoney USD index: empty spot")
            day = datetime.fromtimestamp(int(timestamp), tz=timezone.utc).strftime("%Y-%m-%d")
            return [(day, float(close) / 100.0)]
        except Exception as error:  # noqa: BLE001 - 采集层需要尽量兜底
            spot_errors.append(f"{spot_host}:{type(error).__name__}:{str(error)[:80]}")
            time.sleep(1)
    if last_error is not None:
        raise last_error
    raise RuntimeError("; ".join(spot_errors))


def _collect_usd_vix() -> CollectionTaskResult:
    metrics: list[dict[str, Any]] = []
    errors: list[str] = []
    for series_id, fetcher in (
        ("USD_INDEX", _fetch_usd_index_points),
        ("VIX", _fetch_vix_points),
    ):
        try:
            points = fetcher()
        except Exception as error:
            errors.append(f"{series_id}:{type(error).__name__}:{str(error)[:100]}")
            continue
        definition = macro_series_index()[series_id]
        for day, close in points[-120:]:
            metrics.append(
                _metric(
                    series_id,
                    series_id,
                    day,
                    "DAILY",
                    definition.name,
                    close,
                    definition.definition,
                    definition.calculation_method,
                    metric_key="close",
                )
            )
    plan = PersistPlan(gold_metrics=metrics)
    status = "PASS" if metrics and not errors else "PARTIAL_FAILURE" if metrics else "BLOCKED"
    return _result(
        "USD_INDEX_VIX",
        "美元指数/VIX",
        "eastmoney+cboe",
        status,
        len(metrics),
        f"指标={len(metrics)}；错误={'; '.join(errors) or '无'}",
        "; ".join(errors) or None,
        plan,
    )
